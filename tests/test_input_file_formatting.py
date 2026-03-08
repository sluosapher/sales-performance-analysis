"""Tests for formatting raw input workbooks for browser display."""

from pathlib import Path
import sys

from openpyxl import Workbook

sys.path.insert(0, str(Path(__file__).resolve().parents[1]))

from sales_mcp_server import format_input_file


def test_format_input_file_includes_sheet_headers_and_rows(tmp_path: Path) -> None:
    workbook_path = tmp_path / "raw_data_20260131.xlsx"
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "SalesData"
    sheet.append(["Geo", "FTF_Name", "Revenue ($M)"])
    sheet.append(["NA", "Alice", 12.5])
    workbook.save(workbook_path)

    formatted = format_input_file(workbook_path)

    assert "RAW SALES INPUT DATA" in formatted
    assert "Sheet: SalesData" in formatted
    assert "Geo | FTF_Name | Revenue ($M)" in formatted
    data_line = next(
        line for line in formatted.splitlines() if "NA" in line and "Alice" in line
    )
    assert "12.5" in data_line


def test_format_input_file_respects_max_rows_per_sheet(tmp_path: Path) -> None:
    workbook_path = tmp_path / "raw_data_20260131.xlsx"
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Rows"
    sheet.append(["index"])
    for idx in range(1, 8):
        sheet.append([idx])
    workbook.save(workbook_path)

    formatted = format_input_file(workbook_path, max_rows_per_sheet=3)

    assert "1" in formatted
    assert "2" in formatted
    assert "3" in formatted
    assert "... truncated after 3 rows ..." in formatted
    assert "4" not in formatted


def test_format_input_file_aligns_columns_for_readability(tmp_path: Path) -> None:
    workbook_path = tmp_path / "raw_data_20260131.xlsx"
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Aligned"
    sheet.append(["Geo", "FTF_Name", "Revenue ($M)"])
    sheet.append(["NA", "Al", 1.0])
    sheet.append(["EMEA", "VeryLongSalesName", 1234.56])
    workbook.save(workbook_path)

    formatted = format_input_file(workbook_path)
    lines = formatted.splitlines()

    header_line = next(line for line in lines if "Geo" in line and "FTF_Name" in line)
    row_one = next(line for line in lines if "NA" in line and "Al" in line)
    row_two = next(line for line in lines if "EMEA" in line and "VeryLongSalesName" in line)

    # All rows should have identical width when rendered as fixed-width columns.
    assert len(header_line) == len(row_one) == len(row_two)
