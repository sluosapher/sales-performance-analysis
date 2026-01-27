"""MCP Server for Sales Performance Analysis."""

import asyncio
from mcp.server.fastmcp.server import FastMCP as Server
from mcp.server.stdio import stdio_server
import anyio
from mcp.server.models import InitializationOptions
from mcp.types import Tool
from collections import defaultdict
from typing import Callable, Dict, Iterable, List, NamedTuple, Optional, Tuple
from datetime import date # for quarter_sort_key
from openpyxl import Workbook, load_workbook
from openpyxl.worksheet.worksheet import Worksheet
import re # for QUARTER_PATTERN
from pathlib import Path # Added missing import
from openpyxl.styles import Font # Added missing import

TARGET_GEOS = ["AP", "BRAZIL", "EMEA", "LAS", "MX", "NA"]
REQUIRED_COLUMNS = {
    "Geo", "FTF_Name", "Quarter", "Revenue ($M)", "oh_l3_sub_offering"
}
DEFAULT_ALL_SHEET_NAME = "Top 10 Sales by Geo"
DEFAULT_THINKSHIELD_SHEET_NAME = "Top 10 ThinkShield by Geo"
DEFAULT_TOP_PERCENT_SHEET_NAME = "Top 10% All"
DEFAULT_TOP_PERCENT_SECURITY_SHEET_NAME = "Top 10% Security"
SUMMARY_SHEET_NAME = "Summary"
NUMBER_FORMAT = '[$$-409]#,##0.00'
THINKSHIELD_VALUE = "ThinkShield Security"
THINKSHIELD_VALUE_LOWER = THINKSHIELD_VALUE.lower()
QUARTER_PATTERN = re.compile(r"^FY(?P<year>\d{4})Q(?P<quarter>\d)$")
RAW_DATA_STEM_PATTERN = re.compile(r"^raw_data_(\d{6})$", re.IGNORECASE)

class SalesRow(NamedTuple):
    geo: str
    salesperson: str
    quarter: str
    offering: str
    revenue: float

def to_str(value: object) -> str:
    return str(value).strip() if value is not None else ""

def to_float(value: object) -> float:
    if value in (None, ""):
        return 0.0
    try:
        return float(value)
    except (TypeError, ValueError) as exc:
        raise ValueError(f"Unable to convert {value!r} to float.") from exc

def iter_rows_values(ws: Worksheet, min_row: int = 1, max_rows: int | None = None) -> Iterable[List[object]]:
    """Yield worksheet rows as plain value lists."""
    count = 0
    for row in ws.iter_rows(min_row=min_row, values_only=True):
        yield list(row)
        count += 1
        if max_rows is not None and count >= max_rows:
            break

def build_header_index(header_row: Iterable[object]) -> Dict[str, int]:
    index: Dict[str, int] = {}
    for col_idx, raw_value in enumerate(header_row):
        header = to_str(raw_value)
        if header:
            index[header.strip()] = col_idx
    return index

def quarter_sort_key(value: str, fallback_index: int) -> Tuple[int, int, int]:
    """Return a sort key that orders fiscal quarters chronologically."""
    match = QUARTER_PATTERN.match(value)
    if match:
        year = int(match.group("year"))
        quarter = int(match.group("quarter"))
        return (0, year, quarter)
    return (1, fallback_index, 0)

def load_sales_data(path: Path) -> Tuple[List[str], List[SalesRow]]:
    """Load sales rows from the worksheet that contains the required columns."""
    workbook = load_workbook(filename=path, data_only=True, read_only=True)
    try:
        worksheet: Worksheet | None = None
        header_index: Dict[str, int] | None = None

        for candidate in workbook.worksheets:
            header_row = next(iter_rows_values(candidate, max_rows=1), None)
            if not header_row:
                continue
            candidate_index = build_header_index(header_row)
            if REQUIRED_COLUMNS.issubset(candidate_index):
                worksheet = candidate
                header_index = candidate_index
                break

        if worksheet is None or header_index is None:
            raise ValueError(
                f"No worksheet in {path} contains columns: {', '.join(sorted(REQUIRED_COLUMNS))}"
            )

        quarter_order: Dict[str, int] = {}
        next_order = 0

        rows: List[SalesRow] = []
        for row in iter_rows_values(worksheet, min_row=2):
            geo = to_str(row[header_index["Geo"]])
            if geo not in TARGET_GEOS:
                continue

            sales_person = to_str(row[header_index["FTF_Name"]]) or "blank"
            quarter = to_str(row[header_index["Quarter"]]) or "Unknown"
            revenue = to_float(row[header_index["Revenue ($M)"]])
            offering = to_str(row[header_index["oh_l3_sub_offering"]])

            if quarter not in quarter_order:
                quarter_order[quarter] = next_order
                next_order += 1

            rows.append(SalesRow(geo, sales_person, quarter, offering, revenue))

        quarters = sorted(
            quarter_order.keys(),
            key=lambda value: quarter_sort_key(value, quarter_order[value]),
        )

        return quarters, rows
    finally:
        workbook.close()

def summarize_sales(
    rows: Iterable[SalesRow],
    offering_filter: Callable[[SalesRow], bool] | None = None,
) -> Dict[str, Dict[str, Dict[str, float]]]:
    """Return nested geo -> salesperson -> quarter -> revenue mapping."""
    summary: Dict[str, Dict[str, Dict[str, float]]] = {
        geo: defaultdict(lambda: defaultdict(float)) for geo in TARGET_GEOS
    }
    for entry in rows:
        if offering_filter and not offering_filter(entry):
            continue
        summary[entry.geo][entry.salesperson][entry.quarter] += entry.revenue
    return summary

def sort_salespeople(data: Dict[str, Dict[str, Dict[str, float]]], quarters: List[str]) -> Dict[str, List[Tuple[str, List[float], float]]]:
    """Prepare the top 10 rows per geo with per-quarter totals."""
    result: Dict[str, List[Tuple[str, List[float], float]]] = {}
    for geo in TARGET_GEOS:
        geo_entries = []
        for salesperson, quarter_map in data.get(geo, {}).items():
            quarter_values = [quarter_map.get(q, 0.0) for q in quarters]
            total = sum(quarter_values)
            geo_entries.append((salesperson, quarter_values, total))
        geo_entries.sort(key=lambda item: item[2], reverse=True)
        result[geo] = geo_entries[:10]
    return result

def load_or_create_workbook(output_path: Path) -> Workbook:
    """Load an existing workbook or create a fresh one if it does not exist."""
    if output_path.exists():
        return load_workbook(output_path)
    workbook = Workbook()
    if workbook.sheetnames:
        workbook.remove(workbook[workbook.sheetnames[0]])
    return workbook

def ensure_report_sheet(workbook: Workbook, sheet_name: str) -> Worksheet:
    """Create or replace a worksheet for the report."""
    if sheet_name in workbook.sheetnames:
        del workbook[sheet_name]
    return workbook.create_sheet(sheet_name)

def write_report(
    worksheet: Worksheet,
    quarters: list[str],
    top_sales: dict[str, list[tuple[str, list[float], float]]],
) -> None:
    """Write the geo reports into the worksheet."""
    has_data = any(top_sales.get(geo) for geo in TARGET_GEOS)
    if not has_data:
        worksheet.cell(row=1, column=1, value="No data available.")
        return

    current_row = 1
    header = ["Salesperson"] + quarters + ["Total"]

    for geo in TARGET_GEOS:
        rows = top_sales.get(geo, [])
        if not rows:
            continue

        worksheet.cell(row=current_row, column=1, value=geo)
        current_row += 1

        for col, name in enumerate(header, start=1):
            worksheet.cell(row=current_row, column=col, value=name)
        current_row += 1

        geo_totals = [0.0 for _ in header[1:]]  # quarters + total

        for salesperson, quarter_values, total in rows:
            worksheet.cell(row=current_row, column=1, value=salesperson)
            for offset, value in enumerate(quarter_values, start=2):
                cell = worksheet.cell(row=current_row, column=offset, value=value)
                cell.number_format = NUMBER_FORMAT
                geo_totals[offset - 2] += value
            total_cell = worksheet.cell(row=current_row, column=len(header), value=total)
            total_cell.number_format = NUMBER_FORMAT
            geo_totals[-1] += total
            current_row += 1

        worksheet.cell(row=current_row, column=1, value=f"{geo} Total")
        for offset, value in enumerate(geo_totals, start=2):
            cell = worksheet.cell(row=current_row, column=offset, value=value)
            cell.number_format = NUMBER_FORMAT
        current_row += 2  # blank row between geos

def compute_group_totals(
    summary: Dict[str, Dict[str, Dict[str, float]]],
    quarters: List[str],
    geos: Iterable[str],
) -> Tuple[List[float], float]:
    """Return per-quarter and grand totals for the specified geo collection."""
    quarter_totals = [0.0 for _ in quarters]
    for geo in geos:
        for quarter_map in summary.get(geo, {}).values():
            for idx, quarter in enumerate(quarters):
                quarter_totals[idx] += quarter_map.get(quarter, 0.0)
    grand_total = sum(quarter_totals)
    return quarter_totals, grand_total


def compute_group_top10_totals(
    summary: Dict[str, Dict[str, Dict[str, float]]],
    quarters: List[str],
    geos: Iterable[str],
) -> Tuple[List[float], float]:
    """Aggregate top 10 salesperson revenue by quarter for the geo collection."""
    entries: List[Tuple[List[float], float]] = []
    for geo in geos:
        for quarter_map in summary.get(geo, {}).values():
            quarter_values = [quarter_map.get(q, 0.0) for q in quarters]
            total = sum(quarter_values)
            entries.append((quarter_values, total))

    entries.sort(key=lambda item: item[1], reverse=True)
    top_entries = entries[:10]

    quarter_totals = [0.0 for _ in quarters]
    grand_total = 0.0
    for quarter_values, total in top_entries:
        for idx, value in enumerate(quarter_values):
            quarter_totals[idx] += value
        grand_total += total
    return quarter_totals, grand_total


def write_top_percent_sheet(
    worksheet: Worksheet,
    quarters: List[str],
    summary: Dict[str, Dict[str, Dict[str, float]]],
) -> None:
    """Write the Top 10 percentage tables for each geo group."""
    groups = [
        ("AP", ["AP"]),
        ("EMEA", ["EMEA"]),
        ("NA", ["NA"]),
    ]
    other_geos = [geo for geo in TARGET_GEOS if geo not in {"AP", "EMEA", "NA"}]
    groups.append(("OTHERS", other_geos))

    header = [""] + quarters + ["Grand Total"]
    current_row = 1

    for group_name, geos in groups:
        worksheet.cell(row=current_row, column=1, value=group_name)
        current_row += 1

        for col_idx, label in enumerate(header, start=1):
            worksheet.cell(row=current_row, column=col_idx, value=label)
        current_row += 1

        quarter_totals, grand_total = compute_group_totals(summary, quarters, geos)
        top_quarters, top_grand = compute_group_top10_totals(summary, quarters, geos)

        total_row = ["Sum of Revenue ($M)"] + quarter_totals + [grand_total]
        top_row = ["Top 10 FTF"] + top_quarters + [top_grand]
        percent_values = [
            (top / total) if total else 0.0
            for top, total in zip(top_row[1:], total_row[1:])
        ]
        percent_row = ["Top 10 FTF Rev %"] + percent_values

        for col_idx, value in enumerate(total_row, start=1):
            cell = worksheet.cell(row=current_row, column=col_idx, value=value)
            if col_idx > 1:
                cell.number_format = NUMBER_FORMAT
        current_row += 1

        for col_idx, value in enumerate(top_row, start=1):
            cell = worksheet.cell(row=current_row, column=col_idx, value=value)
            if col_idx > 1:
                cell.number_format = NUMBER_FORMAT
        current_row += 1

        for col_idx, value in enumerate(percent_row, start=1):
            cell = worksheet.cell(row=current_row, column=col_idx, value=value)
            if col_idx > 1:
                cell.number_format = "0%"
        current_row += 2  # blank row between sections

def ensure_summary_sheet(workbook: Workbook, sheet_name: str) -> Worksheet:
    """Create or replace the summary worksheet and place it first."""
    if sheet_name in workbook.sheetnames:
        del workbook[sheet_name]
    return workbook.create_sheet(sheet_name, 0)

def write_summary_sheet(
    worksheet: Worksheet,
    generation_date: str,
    input_filename: str,
    sheet_summaries: List[Tuple[str, str]],
) -> None:
    """Populate the summary worksheet with metadata and tab descriptions."""
    title_cell = worksheet.cell(row=1, column=1, value="Sales Performance Analysis Report")
    title_cell.font = Font(size=24, bold=True)

    generated_cell = worksheet.cell(row=3, column=1, value=f"Generated on: {generation_date}")
    generated_cell.font = Font(size=16, bold=True)

    input_cell = worksheet.cell(row=4, column=1, value=f"Input data: {input_filename}")
    input_cell.font = Font(size=16, bold=True)

    header_tab = worksheet.cell(row=6, column=1, value="Tab")
    header_summary = worksheet.cell(row=6, column=2, value="Summary")
    header_tab.font = Font(size=20, bold=True)
    header_summary.font = Font(size=20, bold=True)

    for index, (sheet_name, description) in enumerate(sheet_summaries, start=7):
        name_cell = worksheet.cell(row=index, column=1, value=sheet_name)
        summary_cell = worksheet.cell(row=index, column=2, value=description)
        name_cell.font = Font(size=20)
        summary_cell.font = Font(size=20)

    worksheet.column_dimensions["A"].width = 28
    worksheet.column_dimensions["B"].width = 70

def extract_timestamp_from_stem(stem: str) -> str | None:
    RAW_DATA_STEM_PATTERN = re.compile(r"^raw_data_(\d{6})$", re.IGNORECASE)
    match = RAW_DATA_STEM_PATTERN.match(stem)
    if match:
        return match.group(1)
    return None

def timestamp_to_date(timestamp: str) -> date:
    if len(timestamp) != 6:
        raise ValueError(f"Timestamp must have 6 digits, got {timestamp!r}.")
    yy = int(timestamp[:2])
    mm = int(timestamp[2:4])
    dd = int(timestamp[4:])
    year = 2000 + yy
    return date(year, mm, dd)

def process_input_file(input_path: Path, output_dir: Path) -> tuple[Path, str]:
    """Process a single input file and generate result."""
    input_dir = input_path.parent
    output_dir.mkdir(parents=True, exist_ok=True)

    timestamp = extract_timestamp_from_stem(input_path.stem)
    if not timestamp:
        raise ValueError(
            f"Input file {input_path.name!r} does not match the expected pattern "
            "'raw_data_YYMMDD.xlsx'."
        )

    quarters, rows = load_sales_data(input_path)
    if not quarters:
        raise ValueError("No quarter data found in the input workbook.")

    output_filename = f"result_{timestamp}.xlsx"
    output_path = output_dir / output_filename

    workbook = load_or_create_workbook(output_path)

    all_sheet = ensure_report_sheet(workbook, DEFAULT_ALL_SHEET_NAME)
    all_summary = summarize_sales(rows)
    all_top_sales = sort_salespeople(all_summary, quarters)
    write_report(all_sheet, quarters, all_top_sales)

    think_sheet = ensure_report_sheet(workbook, DEFAULT_THINKSHIELD_SHEET_NAME)
    think_summary = summarize_sales(
        rows,
        offering_filter=lambda entry: entry.offering.lower() == THINKSHIELD_VALUE_LOWER,
    )
    think_top_sales = sort_salespeople(think_summary, quarters)
    write_report(think_sheet, quarters, think_top_sales)

    top_percent_sheet = ensure_report_sheet(workbook, DEFAULT_TOP_PERCENT_SHEET_NAME)
    write_top_percent_sheet(top_percent_sheet, quarters, all_summary)

    top_percent_security_sheet = ensure_report_sheet(
        workbook, DEFAULT_TOP_PERCENT_SECURITY_SHEET_NAME
    )
    write_top_percent_sheet(top_percent_security_sheet, quarters, think_summary)

    summary_sheet = ensure_summary_sheet(workbook, SUMMARY_SHEET_NAME)
    from datetime import datetime
    generation_date = datetime.now().strftime("%Y-%m-%d")
    sheet_descriptions = [
        (
            DEFAULT_ALL_SHEET_NAME,
            "Top 10 salespeople per geo across all offerings with quarterly and total revenue.",
        ),
        (
            DEFAULT_THINKSHIELD_SHEET_NAME,
            "Top 10 salespeople per geo for the ThinkShield Security offering only.",
        ),
        (
            DEFAULT_TOP_PERCENT_SHEET_NAME,
            "Share of revenue captured by the top 10 sellers versus total sales for AP, EMEA, NA, and OTHERS.",
        ),
        (
            DEFAULT_TOP_PERCENT_SECURITY_SHEET_NAME,
            "ThinkShield Security top 10 revenue share compared to totals for AP, EMEA, NA, and OTHERS.",
        ),
    ]
    write_summary_sheet(summary_sheet, generation_date, input_path.name, sheet_descriptions)

    workbook.save(output_path)
    return output_path, timestamp

app = Server("sales-performance-analysis")

@app.list_tools()
async def handle_list_tools() -> list[Tool]:
    """List available tools."""
    return [
        Tool(
            name="upload-input",
            description="Upload an Excel sales data file for processing",
            inputSchema={
                "type": "object",
                "properties": {
                    "file_name": {
                        "type": "string",
                        "description": "Name of the uploaded file (must match raw_data_YYMMDD.xlsx pattern)"
                    },
                    "content": {
                        "type": "string",
                        "description": "Base64-encoded Excel file content"
                    }
                },
                "required": ["file_name", "content"]
            },
        ),
        Tool(
            name="list-results",
            description="List all available result files",
            inputSchema={
                "type": "object",
                "properties": {},
                "required": []
            },
        ),
        Tool(
            name="get-result",
            description="Get formatted results from a specific result file",
            inputSchema={
                "type": "object",
                "properties": {
                    "file_name": {
                        "type": "string",
                        "description": "Name of the result file (e.g., result_YYMMDD.xlsx)"
                    }
                },
                "required": ["file_name"]
            },
        ),
    ]

@app.call_tool()
async def handle_call_tool(name: str, arguments: dict) -> str:
    """Handle tool execution."""
    import base64
    # from pathlib import Path  # Already imported
    # from tempfile import NamedTemporaryFile # Not needed, direct write

    if name == "upload-input":
        file_name = arguments["file_name"]
        content = arguments["content"]

        if not file_name.endswith(".xlsx"):
            return f"Error: File must be an Excel .xlsx file"

        if not extract_timestamp_from_stem(file_name.replace(".xlsx", "")):
            return f"Error: File name must match pattern raw_data_YYMMDD.xlsx"

        input_dir = Path("input")
        input_dir.mkdir(parents=True, exist_ok=True)

        input_path = input_dir / file_name

        try:
            file_content = base64.b64decode(content)
            with open(input_path, "wb") as f:
                f.write(file_content)
        except Exception as e:
            return f"Error: Failed to save file: {e}"

        try:
            output_dir = Path("output")
            output_path, timestamp = process_input_file(input_path, output_dir)

            return (
                f"File processed successfully!\n"
                f"Input: {file_name}\n"
                f"Output: {output_path.name}\n"
                f"Timestamp: {timestamp}"
            )
        except Exception as e:
            return f"Error: Failed to process file: {e}"

    elif name == "list-results":
        output_dir = Path("output")
        output_dir.mkdir(parents=True, exist_ok=True)

        result_files = sorted(output_dir.glob("result_*.xlsx"))

        if not result_files:
            return "No result files found. Upload and process a file first."

        lines = ["Available Result Files:", "=" * 60]
        for file_path in result_files:
            stat = file_path.stat()
            from datetime import datetime
            mod_time = datetime.fromtimestamp(stat.st_mtime).strftime("%Y-%m-%d %H:%M:%S")
            size_kb = stat.st_size / 1024
            lines.append(f"  {file_path.name}")
            lines.append(f"    Modified: {mod_time}")
            lines.append(f"    Size: {size_kb:.1f} KB")
            lines.append("")

        return "\n".join(lines)

    elif name == "get-result":
        file_name = arguments["file_name"]

        if not file_name.endswith(".xlsx"):
            return "Error: File must be an Excel .xlsx file"

        if not file_name.startswith("result_"):
            return "Error: Result file name must start with 'result_'"

        output_dir = Path("output")
        result_path = output_dir / file_name

        if not result_path.exists():
            available = [f.name for f in output_dir.glob("result_*.xlsx")]
            return f"Error: File {file_name} not found.\nAvailable: {', '.join(available)}"

        try:
            result_text = format_result_file(result_path)
            return result_text
        except Exception as e:
            return f"Error: Failed to format result: {e}"
    # Other tool handlers will go here later

async def main():
    async with stdio_server() as (read_stream, write_stream):
        await app.run(read_stream, write_stream, InitializationOptions())

if __name__ == "__main__":
    anyio.run(main)
