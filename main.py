"""Generate a geo-based top 10 sales report from an input workbook."""

from __future__ import annotations

import argparse
import ctypes
import re
import sys
from collections import defaultdict
from pathlib import Path
from typing import Callable, Dict, Iterable, List, NamedTuple, Tuple

from openpyxl import Workbook, load_workbook
from openpyxl.worksheet.worksheet import Worksheet

TARGET_GEOS = ["AP", "BRAZIL", "EMEA", "LAS", "MX", "NA"]
REQUIRED_COLUMNS = {
    "Geo",
    "FTF_Name",
    "Quarter",
    "Revenue ($M)",
    "oh_l3_sub_offering",
}
DEFAULT_ALL_SHEET_NAME = "Top 10 Sales by Geo"
DEFAULT_THINKSHIELD_SHEET_NAME = "Top 10 ThinkShield by Geo"
DEFAULT_TOP_PERCENT_SHEET_NAME = "Top 10% All"
DEFAULT_TOP_PERCENT_SECURITY_SHEET_NAME = "Top 10% Security"
NUMBER_FORMAT = '[$$-409]#,##0.00'
THINKSHIELD_VALUE = "ThinkShield Security"
THINKSHIELD_VALUE_LOWER = THINKSHIELD_VALUE.lower()
QUARTER_PATTERN = re.compile(r"^FY(?P<year>\d{4})Q(?P<quarter>\d)$")


class SalesRow(NamedTuple):
    geo: str
    salesperson: str
    quarter: str
    offering: str
    revenue: float


class ArgumentError(Exception):
    pass


class PopupArgumentParser(argparse.ArgumentParser):
    def error(self, message: str) -> None:
        raise ArgumentError(message)


def show_error(message: str, title: str = "Sales Report Error") -> None:
    """Display an error message in a Windows popup, with console fallback."""
    try:
        if sys.platform == "win32" and hasattr(ctypes, "windll"):
            ctypes.windll.user32.MessageBoxW(
                None,
                message,
                title,
                0x10,  # MB_ICONERROR
            )
        else:
            print(message, file=sys.stderr)
    except Exception:
        print(message, file=sys.stderr)


def parse_args() -> argparse.Namespace:
    parser = PopupArgumentParser(
        description=(
            "Build a top 10 sales report for each geo group and write the "
            "results to a new worksheet."
        )
    )
    parser.add_argument("input_file", help="Path to the source Excel workbook.")
    parser.add_argument("output_file", help="Path to the Excel workbook where the report will be written.")
    parser.add_argument(
        "--all-sheet-name",
        default=DEFAULT_ALL_SHEET_NAME,
        help=f"Name of the all offerings worksheet (default: {DEFAULT_ALL_SHEET_NAME!r}).",
    )
    parser.add_argument(
        "--thinkshield-sheet-name",
        default=DEFAULT_THINKSHIELD_SHEET_NAME,
        help=f"Name of the ThinkShield worksheet (default: {DEFAULT_THINKSHIELD_SHEET_NAME!r}).",
    )
    parser.add_argument(
        "--top-percent-sheet-name",
        default=DEFAULT_TOP_PERCENT_SHEET_NAME,
        help=f"Name of the Top 10 percent worksheet (default: {DEFAULT_TOP_PERCENT_SHEET_NAME!r}).",
    )
    parser.add_argument(
        "--top-percent-security-sheet-name",
        default=DEFAULT_TOP_PERCENT_SECURITY_SHEET_NAME,
        help=(
            "Name of the ThinkShield Top 10 percent worksheet "
            f"(default: {DEFAULT_TOP_PERCENT_SECURITY_SHEET_NAME!r})."
        ),
    )
    return parser.parse_args()


def load_sales_data(path: Path) -> Tuple[List[str], List[SalesRow]]:
    """Load sales rows from the worksheet that contains the required columns."""
    workbook = load_workbook(filename=path, data_only=True, read_only=True)
    try:
        worksheet: Worksheet | None = None
        header_index: Dict[str, int] | None = None

        for candidate in workbook.worksheets:
            header_row = next(iter_rows_values(candidate, max_rows=1), None)
            if not header_row:
                continue
            candidate_index = build_header_index(header_row)
            if REQUIRED_COLUMNS.issubset(candidate_index):
                worksheet = candidate
                header_index = candidate_index
                break

        if worksheet is None or header_index is None:
            raise ValueError(
                f"No worksheet in {path} contains columns: {', '.join(sorted(REQUIRED_COLUMNS))}"
            )

        quarter_order: Dict[str, int] = {}
        next_order = 0

        rows: List[SalesRow] = []
        for row in iter_rows_values(worksheet, min_row=2):
            geo = to_str(row[header_index["Geo"]])
            if geo not in TARGET_GEOS:
                continue

            sales_person = to_str(row[header_index["FTF_Name"]]) or "blank"
            quarter = to_str(row[header_index["Quarter"]]) or "Unknown"
            revenue = to_float(row[header_index["Revenue ($M)"]])
            offering = to_str(row[header_index["oh_l3_sub_offering"]])

            if quarter not in quarter_order:
                quarter_order[quarter] = next_order
                next_order += 1

            rows.append(SalesRow(geo, sales_person, quarter, offering, revenue))

        quarters = sorted(
            quarter_order.keys(),
            key=lambda value: quarter_sort_key(value, quarter_order[value]),
        )

        return quarters, rows
    finally:
        workbook.close()


def iter_rows_values(ws: Worksheet, min_row: int = 1, max_rows: int | None = None) -> Iterable[List[object]]:
    """Yield worksheet rows as plain value lists."""
    count = 0
    for row in ws.iter_rows(min_row=min_row, values_only=True):
        yield list(row)
        count += 1
        if max_rows is not None and count >= max_rows:
            break


def build_header_index(header_row: Iterable[object]) -> Dict[str, int]:
    index: Dict[str, int] = {}
    for col_idx, raw_value in enumerate(header_row):
        header = to_str(raw_value)
        if header:
            index[header.strip()] = col_idx
    return index


def to_str(value: object) -> str:
    return str(value).strip() if value is not None else ""


def to_float(value: object) -> float:
    if value in (None, ""):
        return 0.0
    try:
        return float(value)
    except (TypeError, ValueError) as exc:
        raise ValueError(f"Unable to convert {value!r} to float.") from exc


def quarter_sort_key(value: str, fallback_index: int) -> Tuple[int, int, int]:
    """Return a sort key that orders fiscal quarters chronologically."""
    match = QUARTER_PATTERN.match(value)
    if match:
        year = int(match.group("year"))
        quarter = int(match.group("quarter"))
        return (0, year, quarter)
    return (1, fallback_index, 0)


def summarize_sales(
    rows: Iterable[SalesRow],
    offering_filter: Callable[[SalesRow], bool] | None = None,
) -> Dict[str, Dict[str, Dict[str, float]]]:
    """Return nested geo -> salesperson -> quarter -> revenue mapping."""
    summary: Dict[str, Dict[str, Dict[str, float]]] = {
        geo: defaultdict(lambda: defaultdict(float)) for geo in TARGET_GEOS
    }
    for entry in rows:
        if offering_filter and not offering_filter(entry):
            continue
        summary[entry.geo][entry.salesperson][entry.quarter] += entry.revenue
    return summary


def sort_salespeople(data: Dict[str, Dict[str, Dict[str, float]]], quarters: List[str]) -> Dict[str, List[Tuple[str, List[float], float]]]:
    """Prepare the top 10 rows per geo with per-quarter totals."""
    result: Dict[str, List[Tuple[str, List[float], float]]] = {}
    for geo in TARGET_GEOS:
        geo_entries = []
        for salesperson, quarter_map in data.get(geo, {}).items():
            quarter_values = [quarter_map.get(q, 0.0) for q in quarters]
            total = sum(quarter_values)
            geo_entries.append((salesperson, quarter_values, total))
        geo_entries.sort(key=lambda item: item[2], reverse=True)
        result[geo] = geo_entries[:10]
    return result


def load_or_create_workbook(output_path: Path) -> Workbook:
    """Load an existing workbook or create a fresh one if it does not exist."""
    if output_path.exists():
        return load_workbook(output_path)

    workbook = Workbook()
    if workbook.sheetnames:
        workbook.remove(workbook[workbook.sheetnames[0]])
    return workbook


def ensure_report_sheet(workbook: Workbook, sheet_name: str) -> Worksheet:
    """Create or replace a worksheet for the report."""
    if sheet_name in workbook.sheetnames:
        del workbook[sheet_name]
    return workbook.create_sheet(sheet_name)


def compute_group_totals(
    summary: Dict[str, Dict[str, Dict[str, float]]],
    quarters: List[str],
    geos: Iterable[str],
) -> Tuple[List[float], float]:
    """Return per-quarter and grand totals for the specified geo collection."""
    quarter_totals = [0.0 for _ in quarters]
    for geo in geos:
        for quarter_map in summary.get(geo, {}).values():
            for idx, quarter in enumerate(quarters):
                quarter_totals[idx] += quarter_map.get(quarter, 0.0)
    grand_total = sum(quarter_totals)
    return quarter_totals, grand_total


def compute_group_top10_totals(
    summary: Dict[str, Dict[str, Dict[str, float]]],
    quarters: List[str],
    geos: Iterable[str],
) -> Tuple[List[float], float]:
    """Aggregate top 10 salesperson revenue by quarter for the geo collection."""
    entries: List[Tuple[List[float], float]] = []
    for geo in geos:
        for quarter_map in summary.get(geo, {}).values():
            quarter_values = [quarter_map.get(q, 0.0) for q in quarters]
            total = sum(quarter_values)
            entries.append((quarter_values, total))

    entries.sort(key=lambda item: item[1], reverse=True)
    top_entries = entries[:10]

    quarter_totals = [0.0 for _ in quarters]
    grand_total = 0.0
    for quarter_values, total in top_entries:
        for idx, value in enumerate(quarter_values):
            quarter_totals[idx] += value
        grand_total += total
    return quarter_totals, grand_total


def write_top_percent_sheet(
    worksheet: Worksheet,
    quarters: List[str],
    summary: Dict[str, Dict[str, Dict[str, float]]],
) -> None:
    """Write the Top 10 percentage tables for each geo group."""
    groups = [
        ("AP", ["AP"]),
        ("EMEA", ["EMEA"]),
        ("NA", ["NA"]),
    ]
    other_geos = [geo for geo in TARGET_GEOS if geo not in {"AP", "EMEA", "NA"}]
    groups.append(("OTHERS", other_geos))

    header = [""] + quarters + ["Grand Total"]
    current_row = 1

    for group_name, geos in groups:
        worksheet.cell(row=current_row, column=1, value=group_name)
        current_row += 1

        for col_idx, label in enumerate(header, start=1):
            worksheet.cell(row=current_row, column=col_idx, value=label)
        current_row += 1

        quarter_totals, grand_total = compute_group_totals(summary, quarters, geos)
        top_quarters, top_grand = compute_group_top10_totals(summary, quarters, geos)

        total_row = ["Sum of Revenue ($M)"] + quarter_totals + [grand_total]
        top_row = ["Top 10 FTF"] + top_quarters + [top_grand]
        percent_values = [
            (top / total) if total else 0.0
            for top, total in zip(top_row[1:], total_row[1:])
        ]
        percent_row = ["Top 10 FTF Rev %"] + percent_values

        for col_idx, value in enumerate(total_row, start=1):
            cell = worksheet.cell(row=current_row, column=col_idx, value=value)
            if col_idx > 1:
                cell.number_format = NUMBER_FORMAT
        current_row += 1

        for col_idx, value in enumerate(top_row, start=1):
            cell = worksheet.cell(row=current_row, column=col_idx, value=value)
            if col_idx > 1:
                cell.number_format = NUMBER_FORMAT
        current_row += 1

        for col_idx, value in enumerate(percent_row, start=1):
            cell = worksheet.cell(row=current_row, column=col_idx, value=value)
            if col_idx > 1:
                cell.number_format = "0%"
        current_row += 2  # blank row between sections


def write_report(
    worksheet: Worksheet,
    quarters: List[str],
    top_sales: Dict[str, List[Tuple[str, List[float], float]]],
) -> None:
    """Write the geo reports into the worksheet."""
    has_data = any(top_sales.get(geo) for geo in TARGET_GEOS)
    if not has_data:
        worksheet.cell(row=1, column=1, value="No data available.")
        return

    current_row = 1
    header = ["Salesperson"] + quarters + ["Total"]

    for geo in TARGET_GEOS:
        rows = top_sales.get(geo, [])
        if not rows:
            continue

        worksheet.cell(row=current_row, column=1, value=geo)
        current_row += 1

        for col, name in enumerate(header, start=1):
            worksheet.cell(row=current_row, column=col, value=name)
        current_row += 1

        geo_totals = [0.0 for _ in header[1:]]  # quarters + total

        for salesperson, quarter_values, total in rows:
            worksheet.cell(row=current_row, column=1, value=salesperson)
            for offset, value in enumerate(quarter_values, start=2):
                cell = worksheet.cell(row=current_row, column=offset, value=value)
                cell.number_format = NUMBER_FORMAT
                geo_totals[offset - 2] += value
            total_cell = worksheet.cell(row=current_row, column=len(header), value=total)
            total_cell.number_format = NUMBER_FORMAT
            geo_totals[-1] += total
            current_row += 1

        # Geo total row
        worksheet.cell(row=current_row, column=1, value=f"{geo} Total")
        for offset, value in enumerate(geo_totals, start=2):
            cell = worksheet.cell(row=current_row, column=offset, value=value)
            cell.number_format = NUMBER_FORMAT
        current_row += 2  # blank row between geos


def main() -> None:
    args = parse_args()
    input_path = Path(args.input_file)
    output_path = Path(args.output_file)

    quarters, rows = load_sales_data(input_path)
    if not quarters:
        raise ValueError("No quarter data found in the input workbook.")

    workbook = load_or_create_workbook(output_path)

    all_sheet = ensure_report_sheet(workbook, args.all_sheet_name)
    all_summary = summarize_sales(rows)
    all_top_sales = sort_salespeople(all_summary, quarters)
    write_report(all_sheet, quarters, all_top_sales)

    think_sheet = ensure_report_sheet(workbook, args.thinkshield_sheet_name)
    think_summary = summarize_sales(
        rows,
        offering_filter=lambda entry: entry.offering.lower() == THINKSHIELD_VALUE_LOWER,
    )
    think_top_sales = sort_salespeople(think_summary, quarters)
    write_report(think_sheet, quarters, think_top_sales)

    top_percent_sheet = ensure_report_sheet(workbook, args.top_percent_sheet_name)
    write_top_percent_sheet(top_percent_sheet, quarters, all_summary)

    top_percent_security_sheet = ensure_report_sheet(
        workbook, args.top_percent_security_sheet_name
    )
    write_top_percent_sheet(top_percent_security_sheet, quarters, think_summary)

    workbook.save(output_path)
    print(
        f"Report written to {output_path} in sheets "
        f"'{args.all_sheet_name}', '{args.thinkshield_sheet_name}', "
        f"'{args.top_percent_sheet_name}', and "
        f"'{args.top_percent_security_sheet_name}'."
    )


if __name__ == "__main__":
    try:
        main()
    except ArgumentError as exc:
        show_error(f"Invalid arguments:\n{exc}")
        sys.exit(2)
    except Exception as exc:
        show_error(f"Unexpected error:\n{exc}")
        sys.exit(1)
